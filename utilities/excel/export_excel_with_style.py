import pathlib

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def export_excel_with_style(df: pd.DataFrame, file_path: pathlib.Path | str):
    # Define the ExcelWriter object
    with pd.ExcelWriter(file_path, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")
        workbook = writer.book
        worksheet = writer.sheets["Sheet1"]

        # Define cell formats for styling
        header_format = workbook.add_format(
            {
                "bold": True,
                "font_color": "white",
                "bg_color": "#007bff",  # Blue color
                "align": "center",
                "valign": "vcenter",
                "border": 1,
            }
        )

        cell_format = workbook.add_format({"align": "center", "valign": "vcenter", "border": 1})

        # Apply header format to the first row (column names)
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)

        # Apply cell format to the data cells
        for row_num in range(1, len(df) + 1):
            for col_num in range(len(df.columns)):
                worksheet.write(row_num, col_num, df.iloc[row_num - 1, col_num], cell_format)

        # Set column widths equal to the maximum content length
        for i, col in enumerate(df.columns):
            max_len = max(df[col].astype(str).str.len().max(), len(col))
            worksheet.set_column(i, i, max_len + 2)

    print("Excel file with styled design created successfully.")


# Função para ajustar automaticamente a largura das colunas
def adjust_column_width(writer, df, sheet_name):
    worksheet = writer.sheets[sheet_name]
    for col_idx, col in enumerate(df.columns, start=1):
        max_length = max(df[col].astype(str).map(len).max(), len(col)) + 4
        worksheet.column_dimensions[get_column_letter(col_idx)].width = max_length


# Função para aplicar filtro à tabela
def apply_filter(writer, df, sheet_name):
    worksheet = writer.sheets[sheet_name]
    table_ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
    table = Table(displayName=sheet_name.replace(" ", "_"), ref=table_ref)
    style = TableStyleInfo(
        # name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    worksheet.add_table(table)
